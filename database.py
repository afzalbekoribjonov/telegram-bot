import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from config import DEFAULT_LANGUAGE, SUPER_ADMIN_ID  # Ensure SUPER_ADMIN_ID is correctly defined in config.py

DB_NAME = "data.db"


# -------------------- INIT --------------------
def init_db():
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                first_name TEXT,
                last_name TEXT,
                username TEXT,
                phone TEXT,
                language TEXT DEFAULT 'uz',
                balance INTEGER DEFAULT 0,
                user_type TEXT DEFAULT 'oddiy'
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS admins (
                user_id INTEGER PRIMARY KEY
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS prices (
                diamond_name TEXT PRIMARY KEY,
                price_uzs INTEGER,
                price_rub INTEGER
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS channels (
                username TEXT PRIMARY KEY,
                name TEXT,
                link TEXT
            )
        """)


# -------------------- USER --------------------
def add_user(user, phone=None):
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT 1 FROM users WHERE user_id = ?", (user.id,))
        if not cursor.fetchone():
            cursor.execute("""
                INSERT INTO users (user_id, first_name, last_name, username, phone, language)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                user.id,
                user.first_name or '',
                user.last_name or '',
                user.username or '',
                phone,
                DEFAULT_LANGUAGE
            ))


def set_user_language(user_id, lang_code):
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET language = ? WHERE user_id = ?", (lang_code, user_id))


def get_user_language(user_id):
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT language FROM users WHERE user_id = ?", (user_id,))
        result = cursor.fetchone()
        return result[0] if result else DEFAULT_LANGUAGE


def get_all_users():
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT user_id, first_name, last_name, username, phone, balance, user_type FROM users")
        return cursor.fetchall()


def update_user_phone(user_id, phone):
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET phone = ? WHERE user_id = ?", (phone, user_id))


def get_user_balance(user_id):
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT balance FROM users WHERE user_id = ?", (user_id,))
        result = cursor.fetchone()
        return result[0] if result else 0


def get_user_type(user_id):
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT user_type FROM users WHERE user_id = ?", (user_id,))
        result = cursor.fetchone()
        return result[0] if result else "oddiy"


def add_diamonds_to_user(user_id: int, amount: int):
    """
    Foydalanuvchining hisobiga olmos qo'shadi.
    Agar foydalanuvchi bazada mavjud bo'lmasa, uni avval yaratadi.
    """
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()

        # Foydalanuvchi mavjudligini tekshirish
        cursor.execute("SELECT balance FROM users WHERE user_id = ?", (user_id,))
        user_balance = cursor.fetchone()

        if user_balance:
            # Foydalanuvchi mavjud bo'lsa, balansini yangilash
            new_balance = user_balance[0] + amount
            cursor.execute("UPDATE users SET balance = ? WHERE user_id = ?", (new_balance, user_id))
        else:
            # Foydalanuvchi mavjud bo'lmasa, uni yaratish va balansni o'rnatish
            cursor.execute("""
                INSERT INTO users (user_id, balance)
                VALUES (?, ?)
            """, (user_id, amount))


# -------------------- ADMIN --------------------
def add_admin(user_id: int) -> bool:
    """Adds a user ID to the admins table. Returns True if added, False if already exists."""
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        # Check if user already exists as admin
        cursor.execute("SELECT 1 FROM admins WHERE user_id = ?", (user_id,))
        if cursor.fetchone():
            return False  # Already an admin
        cursor.execute("INSERT INTO admins (user_id) VALUES (?)", (user_id,))
        return True  # Successfully added


def remove_admin(user_id: int) -> bool:
    """
    Removes a user ID from the admins table.
    Returns True if removed, False if not an admin or if trying to remove SUPER_ADMIN_ID.
    """
    if user_id == SUPER_ADMIN_ID:
        # Prevent removal of the super admin
        print(f"Attempted to remove SUPER_ADMIN_ID: {user_id}. Operation denied.")
        return False

    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM admins WHERE user_id = ?", (user_id,))
        # Check if any row was actually deleted
        return cursor.rowcount > 0


def is_admin(user_id: int) -> bool:
    if user_id == SUPER_ADMIN_ID:
        return True
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT 1 FROM admins WHERE user_id = ?", (user_id,))
        return bool(cursor.fetchone())


def get_all_admins():
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT user_id FROM admins")
        return [row[0] for row in cursor.fetchall()]


# -------------------- PRICES --------------------
def set_price(diamond_name: str, price_uzs: int, price_rub: int):
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO prices (diamond_name, price_uzs, price_rub)
            VALUES (?, ?, ?)
            ON CONFLICT(diamond_name) DO UPDATE SET
                price_uzs = excluded.price_uzs,
                price_rub = excluded.price_rub
        """, (diamond_name, price_uzs, price_rub))


def get_all_prices():
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT diamond_name, price_uzs, price_rub FROM prices")
        return cursor.fetchall()


# -------------------- EXPORT --------------------
def export_users_to_excel(file_path='users.xlsx'):
    users = get_all_users()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    headers = ["ID", "First Name", "Last Name", "Username", "Phone", "Balance (so'm)", "User Type"]
    ws.append(headers)

    style_header(ws, headers)
    fill_users(ws, users)

    wb.save(file_path)
    return file_path


# Excel yordamchi funksiyalar
def style_header(ws, headers):
    font = Font(bold=True, size=12)
    align = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col_num, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=title)
        cell.font = font
        cell.alignment = align
        cell.border = border


def fill_users(ws, users):
    align = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for row_idx, row in enumerate(users, start=2):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = align
            cell.border = border

    # Automatik ustun kengligi
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max_len + 4